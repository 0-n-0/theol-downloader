import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

import numpy as np
import requests
import os
import urllib.parse
import codecs

from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
import concurrent
from urllib.parse import unquote


# ...其他导入和函数定义...
hex_to_letter = {
    '0': 'A',
    '1': 'B',
    '2': 'C',
    '3': 'D',
    '4': 'E',
    '5': 'F',
    '6': 'G',
    '7': 'H',
    '8': 'I',
    '9': 'J',
    'a': 'K',
    'b': 'L',
    'c': 'M',
    'd': 'N',
    'e': 'O',
    'f': 'P'
}

letter_to_hex = {
    'A': '0',
    'B': '1',
    'C': '2',
    'D': '3',
    'E': '4',
    'F': '5',
    'G': '6',
    'H': '7',
    'I': '8',
    'J': '9',
    'K': 'a',
    'L': 'b',
    'M': 'c',
    'N': 'd',
    'O': 'e',
    'P': 'f'
}

def encrypt():
    content = input_text.get("1.0", tk.END).strip()
    encoded_content = encode(content)
    output_text.delete("1.0", tk.END)
    output_text.insert("1.0", encoded_content)

def decrypt():
    content = input_text.get("1.0", tk.END).strip()
    decoded_content = decode(content)
    output_text.delete("1.0", tk.END)
    output_text.insert("1.0", decoded_content)

def encode(content):
    # 对内容进行gb2312编码
    encoded_content = content.encode('gb2312')

    # 对编码后的内容进行hex编码
    hex_encoded_content = codecs.encode(encoded_content, 'hex')

    # 将结果的十六进制字符转换为对应的字母
    letter_encoded_content = ''.join(hex_to_letter[c] for c in hex_encoded_content.decode())

    return letter_encoded_content


def encode_id(id):
    # 对内容进行gb2312编码
    encoded_content = id.encode('gb2312')

    # 对编码后的内容进行hex编码
    hex_encoded_content = codecs.encode(encoded_content, 'hex')

    # 将结果的十六进制字符转换为对应的字母
    letter_encoded_content = ''.join(hex_to_letter[c] for c in hex_encoded_content.decode())

    return letter_encoded_content

def decode(content):
    hex_encoded_content = ''.join(letter_to_hex[c] for c in content)
    encoded_content = codecs.decode(hex_encoded_content, 'hex')
    decoded_content = encoded_content.decode('gb2312')
    return decoded_content

def check_url(url):
    response = None
    try:
        response = requests.get(url, stream=True, timeout=10)
        if 'Content-Type' in response.headers:
            return True
    except Exception as e:
        pass
    finally:
        if response is not None:
            response.close()
    return False

def generate_urls_analysis(start_id, end_id, url_keyword, step):
    return [f"{url_keyword}/meol/common/ckeditor/openfile.jsp?id={encode_id(str(id).zfill(6))}" for id in
            range(start_id, end_id + 1, step)]

def start_checking(start_id, end_id, step, url_keyword, thread_count, bins):
    urls = generate_urls_analysis(start_id, end_id, url_keyword, step)
    valid_ids = []
    with ThreadPoolExecutor(max_workers=thread_count) as executor:
        for url, is_valid in tqdm(zip(urls, executor.map(check_url, urls)), total=len(urls)):
            id = decode(url.split('=')[-1])
            if is_valid:
                valid_ids.append(int(id))

    valid_ids_np = np.array(valid_ids)
    bins_np = np.linspace(start_id, end_id, bins + 1)
    counts, _ = np.histogram(valid_ids_np, bins=bins_np)
    attempts_per_bin = (end_id - start_id) // bins // step
    valid_ratios = counts / attempts_per_bin

    plt.figure(figsize=(int(bins / 5), 6))
    plt.bar(range(1, bins + 1), valid_ratios)
    plt.title('Distribution of valid IDs')
    plt.xlabel('ID Range')
    plt.ylabel('Valid Ratio')
    plt.xticks(range(1, bins + 1), labels=[f'{bins_np[i]:.0f}-{bins_np[i + 1]:.0f}' for i in range(bins)], rotation=90)
    plt.tight_layout()
    plt.savefig(os.path.join(os.getcwd(), f"ID_Distribution_{start_id}_{end_id}.png"))
    plt.close()
    print(f"The distribution graph is saved as 'ID_Distribution_{start_id}_{end_id}.png' in the current directory.")

def create_application_frame(notebook):
    app_frame = ttk.Frame(notebook)
    notebook.add(app_frame, text="ID分布探测器")

    # Adding widgets to the frame
    tk.Label(app_frame, text="起始 ID").grid(row=0, column=0)
    start_id_entry = tk.Entry(app_frame)
    start_id_entry.grid(row=0, column=1)

    tk.Label(app_frame, text="结束 ID").grid(row=1, column=0)
    end_id_entry = tk.Entry(app_frame)
    end_id_entry.grid(row=1, column=1)

    tk.Label(app_frame, text="步幅").grid(row=2, column=0)
    step_entry = tk.Entry(app_frame)
    step_entry.grid(row=2, column=1)

    tk.Label(app_frame, text="URL 关键字").grid(row=3, column=0)
    url_keyword_entry = tk.Entry(app_frame)
    url_keyword_entry.grid(row=3, column=1)

    tk.Label(app_frame, text="线程数").grid(row=4, column=0)
    thread_count_entry = tk.Entry(app_frame)
    thread_count_entry.grid(row=4, column=1)

    tk.Label(app_frame, text="柱状图柱数").grid(row=5, column=0)
    bins_entry = tk.Entry(app_frame)
    bins_entry.grid(row=5, column=1)

    start_button = tk.Button(app_frame, text="开始探测")
    start_button["command"] = lambda: start_checking(
        int(start_id_entry.get()),
        int(end_id_entry.get()),
        int(step_entry.get()),
        url_keyword_entry.get(),
        int(thread_count_entry.get()),
        int(bins_entry.get())
    )
    start_button.grid(row=6, column=0, columnspan=2)

    return app_frame

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.start_id_label = tk.Label(self, text="Start ID")
        self.start_id_label.pack()
        self.start_id_entry = tk.Entry(self)
        self.start_id_entry.pack()

        self.end_id_label = tk.Label(self, text="End ID")
        self.end_id_label.pack()
        self.end_id_entry = tk.Entry(self)
        self.end_id_entry.pack()

        self.step_label = tk.Label(self, text="Step")
        self.step_label.pack()
        self.step_entry = tk.Entry(self)
        self.step_entry.pack()

        self.url_keyword_label = tk.Label(self, text="URL Keyword")
        self.url_keyword_label.pack()
        self.url_keyword_entry = tk.Entry(self)
        self.url_keyword_entry.pack()

        self.thread_count_label = tk.Label(self, text="Thread Count")
        self.thread_count_label.pack()
        self.thread_count_entry = tk.Entry(self)
        self.thread_count_entry.pack()

        self.bins_label = tk.Label(self, text="Bins")
        self.bins_label.pack()
        self.bins_entry = tk.Entry(self)
        self.bins_entry.pack()

        self.start_button = tk.Button(self)
        self.start_button["text"] = "Start"
        self.start_button["command"] = self.start_checking
        self.start_button.pack()

        self.quit = tk.Button(self, text="QUIT", fg="red", command=self.master.destroy)
        self.quit.pack()

    def start_checking(self):
        try:
            start_id = int(self.start_id_entry.get())
            end_id = int(self.end_id_entry.get())
            step = int(self.step_entry.get())
            url_keyword = self.url_keyword_entry.get()
            thread_count = int(self.thread_count_entry.get())
            bins = int(self.bins_entry.get())
            if start_id > end_id or step <= 0 or thread_count <= 0 or bins <= 0:
                raise ValueError
            start_checking(start_id, end_id, step, url_keyword, thread_count, bins)
        except ValueError:
            print("Invalid input! Please ensure all inputs are positive integers and start_id is less than end_id.")

def clean_filename(filename):
    return ''.join(c for c in filename if c.isprintable() and c not in ['*', ':', '/', '\\', '?', '[', ']'])


def get_file_info(url):
    response = None
    try:
        response = requests.get(url, stream=True)
        content_disposition = response.headers.get('Content-Disposition')
        content_length = response.headers.get('Content-Length')

        # 解密 URL 后面的 ID
        id = decode(url.split('=')[-1])

        if content_disposition and content_length:
            file_name = content_disposition.split('filename=')[-1].strip('"')
            file_name = urllib.parse.unquote(file_name)
            file_name = clean_filename(file_name)
            file_size_bytes = int(content_length)
            file_size_mb = round(file_size_bytes / (1024 * 1024), 2)
            return [id, file_name, file_size_bytes, file_size_mb, url]
    except Exception as e:
        print(f"无效的 URL {url}: {e}")
    finally:
        if response is not None:
            response.close()


def generate_urls(start_id, end_id, url_keyword):
    return [f"{url_keyword}/meol/common/ckeditor/openfile.jsp?id={encode_id(str(id).zfill(6))}" for id in
            range(start_id, end_id + 1)]


def generate_file_name(extension):
    base_name = "file_info_"
    suffix = 1
    while os.path.exists(f"{base_name}{suffix}.{extension}"):
        suffix += 1
    return f"{base_name}{suffix}.{extension}"


def start_processing(start_id, end_id, num_threads, url_keyword, sort_by_size):
    urls = generate_urls(start_id, end_id, url_keyword)

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        results = list(tqdm(executor.map(get_file_info, urls), total=len(urls)))

    results = [result for result in results if result]

    # 按文件大小排序
    if sort_by_size:
        results.sort(key=lambda x: x[2], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "文件名", "size (bit)", "size (MB)", "URL"])

    for result in results:
        ws.append(result)

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 30 else 30
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    file_name = generate_file_name("xlsx")
    wb.save(file_name)
    print(f"文件信息收集完成，结果保存在程序目录下的： {file_name}")

def download_file(url, directory):
    response = requests.get(url, stream=True)

    if response.status_code != 200:
        print(f"无效 URL: {url}")
        return

    content_disposition = response.headers.get('Content-Disposition')
    if content_disposition and 'filename=' in content_disposition:
        file_name = content_disposition.split('filename=')[-1].strip('"')
    else:
        return

    file_name = unquote(file_name)

    original_file_name = file_name
    i = 1
    while os.path.exists(os.path.join(directory, file_name)):
        file_name = f"{original_file_name}_{i}"
        i += 1

    with open(os.path.join(directory, file_name), 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

def get_urls_from_xlsx(file, start_row, end_row):
    workbook = load_workbook(filename=file)
    worksheet = workbook.active

    urls = []
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row):
        urls.append(row[4].value)  # 假设URL在第四列

    return urls

def start_downloading(file, start_row, end_row, directory, num_threads):
    urls = get_urls_from_xlsx(file, start_row, end_row)
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        with tqdm(total=len(urls), dynamic_ncols=True) as pbar:
            futures = {executor.submit(download_file, url, directory): url for url in urls}
            for future in concurrent.futures.as_completed(futures):
                pbar.update()


root = tk.Tk()
root.title("T001_THEOL")
root.geometry('370x270')

# 创建Notebook
notebook = ttk.Notebook(root)

# 创建新的标签页
# 创建并显示新的GUI
create_application_frame(notebook)
# 创建新的标签页
collector_frame = ttk.Frame(notebook)
notebook.add(collector_frame, text="文件信息收集器")

# 在文件信息收集器的标签页中添加控件
# 以下是你原有的文件信息收集器的GUI代码，稍作修改以适应新的界面
tk.Label(collector_frame, text="起始 ID").grid(row=0)
tk.Label(collector_frame, text="结束 ID").grid(row=1)
tk.Label(collector_frame, text="线程数").grid(row=2)
tk.Label(collector_frame, text="URL 关键字").grid(row=3)

start_id = tk.Entry(collector_frame)
end_id = tk.Entry(collector_frame)
num_threads = tk.Entry(collector_frame)
url_keyword = tk.Entry(collector_frame)

start_id.grid(row=0, column=1)
end_id.grid(row=1, column=1)
num_threads.grid(row=2, column=1)
url_keyword.grid(row=3, column=1)

tk.Label(collector_frame, text="格式：'http://XXX.XXX.XXX'\n(或https......)").grid(row=4, column=1)

sort_by_size = tk.BooleanVar()
tk.Checkbutton(collector_frame, text="按大小排序", variable=sort_by_size).grid(row=5, column=0, columnspan=2)

tk.Button(collector_frame, text="开始收集",
          command=lambda: start_processing(int(start_id.get()), int(end_id.get()), int(num_threads.get()),
                                           url_keyword.get(), sort_by_size.get())).grid(row=6, column=0, columnspan=2)

# 创建批量下载器的标签页
downloader_frame = ttk.Frame(notebook)
notebook.add(downloader_frame, text="批量下载器")

# 在批量下载器的标签页中添加控件
# 以下是你原有的批量下载器的GUI代码，稍作修改以适应新的界面
file_label = tk.Label(downloader_frame, text="选择字典表格:")
file_label.grid(row=0, column=0, sticky='e')
file_entry = tk.Entry(downloader_frame, width=20)
file_entry.grid(row=0, column=1)
file_button = tk.Button(downloader_frame, text="选择", command=lambda: file_entry.insert(0, filedialog.askopenfilename()))
file_button.grid(row=0, column=2)

start_label = tk.Label(downloader_frame, text="下载起始行:")
start_label.grid(row=1, column=0, sticky='e')
start_entry = tk.Entry(downloader_frame)
start_entry.grid(row=1, column=1)

tk.Label(downloader_frame, text="最小为2").grid(row=1, column=2)


end_label = tk.Label(downloader_frame, text="下载结束行:")
end_label.grid(row=2, column=0, sticky='e')
end_entry = tk.Entry(downloader_frame)
end_entry.grid(row=2, column=1)

directory_label = tk.Label(downloader_frame, text="保存目录:")
directory_label.grid(row=3, column=0, sticky='e')
directory_entry = tk.Entry(downloader_frame, width=20)
directory_entry.grid(row=3, column=1)
directory_button = tk.Button(downloader_frame, text="选择", command=lambda: directory_entry.insert(0, filedialog.askdirectory()))
directory_button.grid(row=3, column=2)

thread_label = tk.Label(downloader_frame, text="线程数:")
thread_label.grid(row=4, column=0, sticky='e')
thread_entry = tk.Entry(downloader_frame)
thread_entry.grid(row=4, column=1)

start_button = tk.Button(downloader_frame, text="开始下载",
                         command=lambda: start_downloading(file_entry.get(), int(start_entry.get()), int(end_entry.get()), directory_entry.get(), int(thread_entry.get())))
start_button.grid(row=5, column=0, columnspan=3)

# 创建新的标签页
# 创建并显示新的GUI
# 创建新的标签页
collector_frame = ttk.Frame(notebook)
notebook.add(collector_frame, text="编码器")
# 创建输入框和标签
input_label = ttk.Label(collector_frame, text="输入:")
input_label.pack()
input_text = tk.Text(collector_frame, height=5)  # 设置输入框的大小
input_text.pack()

# 创建输出框和标签
output_label = ttk.Label(collector_frame, text="输出:")
output_label.pack()
output_text = tk.Text(collector_frame, height=5)  # 设置输出框的大小
output_text.pack()

# 创建按钮
encrypt_button = ttk.Button(collector_frame, text="加密", command=encrypt)
encrypt_button.pack()
decrypt_button = ttk.Button(collector_frame, text="解密", command=decrypt)
decrypt_button.pack()

# 创建新的标签页
# 创建并显示新的GUI
# 创建新的标签页
collector_frame = ttk.Frame(notebook)
notebook.add(collector_frame, text="单文件下载器")



notebook.pack(expand=1, fill="both")  # 将notebook添加到root窗口

root.mainloop()
