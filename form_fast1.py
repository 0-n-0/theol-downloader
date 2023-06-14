import codecs

import requests
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from tqdm import tqdm
import os
import urllib.parse
from openpyxl.utils import get_column_letter

# 创建字母和十六进制数之间的映射

hex_to_letter = {
    '0': 'A',
    '1': 'B',
    '2': 'C',
    '3': 'D',
    '4': 'E',
    '5': 'F',
    '6': 'G',
    '7': 'H',
    '8': 'I',
    '9': 'J',
    'a': 'K',
    'b': 'L',
    'c': 'M',
    'd': 'N',
    'e': 'O',
    'f': 'P'
}


def encode_id(id):
    # 对内容进行gb2312编码
    encoded_content = id.encode('gb2312')

    # 对编码后的内容进行hex编码
    hex_encoded_content = codecs.encode(encoded_content, 'hex')

    # 将结果的十六进制字符转换为对应的字母
    letter_encoded_content = ''.join(hex_to_letter[c] for c in hex_encoded_content.decode())

    return letter_encoded_content


def clean_filename(filename):
    return ''.join(c for c in filename if c.isprintable() and c not in ['*', ':', '/', '\\', '?', '[', ']'])


def get_file_info(url):
    response = None
    try:
        response = requests.get(url, stream=True)
        content_disposition = response.headers.get('Content-Disposition')
        content_length = response.headers.get('Content-Length')

        if content_disposition and content_length:
            file_name = content_disposition.split('filename=')[-1].strip('"')
            file_name = urllib.parse.unquote(file_name)
            file_name = clean_filename(file_name)  # 清理文件名
            file_size_bytes = int(content_length)
            file_size_mb = round(file_size_bytes / (1024 * 1024), 2)
            return [file_name, file_size_bytes, file_size_mb, url]
    except Exception as e:
        print(f"无效的 URL {url}: {e}")
    finally:
        if response is not None:
            response.close()


def generate_urls(start_id, end_id, url_keyword):
    return [f"{url_keyword}/meol/common/ckeditor/openfile.jsp?id={encode_id(str(id).zfill(6))}" for id in
            range(start_id, end_id + 1)]


def generate_file_name(extension):
    base_name = "file_info_"
    suffix = 1
    while os.path.exists(f"{base_name}{suffix}.{extension}"):
        suffix += 1
    return f"{base_name}{suffix}.{extension}"


def start_processing(start_id, end_id, num_threads, url_keyword):
    urls = generate_urls(start_id, end_id, url_keyword)

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        results = list(tqdm(executor.map(get_file_info, urls), total=len(urls)))

    # 删除空结果
    results = [result for result in results if result]

    # 保存结果到 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.append(["文件名", "size (bit)", "size (MB)", "URL"])

    # 按文件大小排序
    results.sort(key=lambda x: x[1], reverse=True)

    for result in results:
        ws.append(result)

    # 设置列宽
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]  # Convert to list
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 30 else 30  # 设置最大宽度为30
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    file_name = generate_file_name("xlsx")
    wb.save(file_name)
    print(f"文件信息收集完成，结果保存在程序目录下的： {file_name}")


def start_gui():
    root = tk.Tk()
    root.title("文件信息收集器")
    root.geometry('256x128')

    tk.Label(root, text="起始ID").grid(row=0)
    tk.Label(root, text="结束ID").grid(row=1)
    tk.Label(root, text="线程数").grid(row=2)
    tk.Label(root, text="URL关键字").grid(row=3)

    start_id = tk.Entry(root)
    end_id = tk.Entry(root)
    num_threads = tk.Entry(root)
    url_keyword = tk.Entry(root)

    start_id.grid(row=0, column=1)
    end_id.grid(row=1, column=1)
    num_threads.grid(row=2, column=1)
    url_keyword.grid(row=3, column=1)

    tk.Button(root, text="开始",
              command=lambda: start_processing(int(start_id.get()), int(end_id.get()), int(num_threads.get()),
                                               url_keyword.get())).grid(row=4, column=0, columnspan=2)

    root.mainloop()


if __name__ == "__main__":
    start_gui()
